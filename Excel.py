# -*- coding: utf-8 -*-
import openpyxl
import sys
import win32api
from typing import List

class Excel:
    __employees_sheet: openpyxl.Workbook
    __rangeMax_row: []

    def __init__(self) -> None:
        self.__employees_sheet = self.employees_sheet()
        self.__rangeMax_row = self.rangeMax_row()

    def employees_sheet(self) -> openpyxl.Workbook:
        try:
            filename = sys.argv[1]
            excel_file = openpyxl.load_workbook(filename)
            return excel_file.active
        except:
            win32api.MessageBox(0, 'Прием на обучение', 'Ожидается файл')
            
    def rangeMax_row(self) -> List[int]:
        result = []
        for x in range(1, self.__employees_sheet.max_row+1):
            result.append(x)
        return result
    
    def getEmployeesSheet(self) -> openpyxl.Workbook:
        return self.__employees_sheet

    def getCell(self, i: int, j: int) -> str:
        result = self.__employees_sheet.cell(row=i, column=j).value
        return result

