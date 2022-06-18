# -*- coding: utf-8 -*-
from requests import patch
from CompetitionGroup import CompetitionGroup
import openpyxl
import win32api

class CheckList:
    __c: CompetitionGroup
    
    def __init__(self) -> None:
        self.__c = CompetitionGroup()
    
    def __lineExist(self, line, a, x, employees_sheet) -> bool:
        result: bool = (( line == None ) and (len(a.abiturient() ) > 0)) or (x == employees_sheet.max_row)
        return result

    def create(self, line, a, x, employees_sheet) -> None:
        self.__c.setName(line)
        if  self.__lineExist(line, a, x, employees_sheet):
            count = 0
            print( len(a.abiturient()) )
            print( '\n' )
            count = len(a.abiturient())    
            for i in range( count ):   
                one = [None]*46
                temp = [None]*32
                for y in range(employees_sheet.max_column):
                    temp[y] = a.abiturient()[i-1][y]
                one[0] = temp[1] #ФИО 1-> 1
                one[2] = temp[30] #Уникальный код 23 -> 30
                if "Среднее специальное" in temp[24]: one[3] = 3
                if "Начальное профессиональное" in temp[24]: one[3] = 3
                if "Среднее общее образование" in temp[24]: one[3] = 2
                one[11] = temp[17] #Номер заявления 11 -> 17
                if "Копия" in temp[12]: one[14] = 0
                if "Оригинал" in temp[12]: one[14] = 1
                if temp[15] == None: one[15] = 0
                if temp[15] != None: one[15] = 1
                if temp[16] == None: one[16] = 0
                if temp[16] != None: one[16] = 1
                one[17] = temp[18]
                one[18] = 1
                if temp[26] != None: one[18] = 2
                if temp[27] != None: one[18] = 3
                if temp[28] != None: one[18] = 4
                if temp[29] != None: one[18] = 4
                if "Заочная" in temp[19]: one[19] = 3
                if "Очная" in temp[19]: one[19] = 1
                if "Очно-заочная" in temp[19]: one[19] = 2
                if "Федеральный бюджет" in temp[20]: one[20] = 1
                if "Внебюджетные средства" in temp[20]: one[20] = 5
                one[21] = 1
                one[26] = 1 
                if temp[5] != None: 
                    one[28] = temp[5] # Математика
                if "Е" in temp[6]: one[29] = 1
                if "Э" in temp[6]: one[29] = 2
                if temp[7] != None: 
                    one[31] = temp[7] # Информатика и ИКТ / Физика
                if "Е" in temp[8]: one[32] = 1
                if "Э" in temp[8]: one[32] = 2
                if temp[9] != None: 
                    one[34] = temp[9] # Русский язык
                if "Е" in temp[10]: one[35] = 1
                if "Э" in temp[10]: one[35] = 2

                try:
                    xlsx: str = '.xlsx'                
                    if i == 0:
                        path = r"c:\Program Files (x86)\Vikon1C\template_abiturs.xlsx"
                        template_file = openpyxl.load_workbook(path)
                    else:
                        template_file = openpyxl.load_workbook(self.__c.getName() + xlsx)
                    template_sheet = template_file['перечень абитуриентов']
                    template_sheet.append(one)
                    template_file.save(self.__c.getName() + xlsx)
                except:
                    win32api.MessageBox(0, 'template_abiturs.xlsx/nперечень абитуриентов', 'Ожидается файл')

            for i in range( count):
                willDel = a.abiturient().pop()
